import os
import re
import clip
import torch
import numpy as np
from PIL import Image
from openpyxl import load_workbook

device = "cuda" if torch.cuda.is_available() else "cpu"
model, preprocess = clip.load('ViT-L/14', device=device)

def forward_modality(model, preprocess, data, flag):
    device = next(model.parameters()).device
    if flag == 'img':
        data = preprocess(data).unsqueeze(0)
        features = model.encode_image(data.to(device))
    elif flag == 'txt':
        data = clip.tokenize(data)
        features = model.encode_text(data.to(device))
    else:
        raise TypeError
    # print(flag, features.shape)
    return features

@torch.no_grad()
def calculate_clip_score(model, preprocess, first_data, second_data, first_flag='txt', second_flag='img'):
    first_features = forward_modality(model, preprocess, first_data, first_flag)
    second_features = forward_modality(model, preprocess, second_data,second_flag)

    # normalize features
    first_features = first_features / first_features.norm(dim=1, keepdim=True).to(torch.float32)
    second_features = second_features / second_features.norm(dim=1, keepdim=True).to(torch.float32)

    # calculate scores
    # score = logit_scale * (second_features * first_features).sum()
    score = (second_features * first_features).sum()
    return score

def compute_image_text_alignment(image_path, text):
    global model, preprocess

    image = Image.open(image_path)
    score = calculate_clip_score(model, preprocess, first_data=text, second_data=image, first_flag='txt', second_flag='img').cpu().numpy()
    return score

def read_excel(file_path):
    workbook = load_workbook(file_path)
    return workbook

def write_excel(file_path, workbook):
    workbook.save(file_path)

def process_images_and_update_excel(folder_path, excel_path, prompt):
    prompts = prompt
    process_percentage = 0
    workbook = read_excel(excel_path)

    seed_pattern = re.compile(r'seed(\d+)')
    prompt_type_pattern = re.compile(r'(personal_prompt|public_prompt)')

    seed_match = seed_pattern.search(folder_path)
    prompt_type_match = prompt_type_pattern.search(folder_path)

    if seed_match and prompt_type_match:
        seed = seed_match.group(1)
        prompt_type = prompt_type_match.group(1)
    else:
        raise ValueError("Unable to find seed and prompt type in folder path.")

    for filename in os.listdir(folder_path):
        if filename.endswith("sharing.png"):
            process_percentage += 1
            parts = filename.split('_')
            if len(parts) == 4:
                x, y, z, _ = parts
                x, y, z = int(x), int(y), int(z)
                if 0 <= x < len(prompts) and 0 <= y < len(prompts):
                    image_path = os.path.join(folder_path, filename)
                    text = prompts[y]
                    similarity = compute_image_text_alignment(image_path, text)
                    similarity_value = float(similarity)
                    sheet_name = f"seed{seed}_{prompt_type}{x}"
                    sheet = workbook[sheet_name]
                    sheet.cell(row=y+2, column=z+2).value = similarity_value
                    print(f"{process_percentage} processed")
                    write_excel(excel_path, workbook)
                else:
                    print(f"File: {filename} has an index out of the range of the prompts list.")
            else:
                print(f"File: {filename} has an incorrect format.")

personal_prompts = [
    "A cat is sleeping peacefully on a sunlit window sill.",
    "A cat is playing with a ball of yarn in a cozy living room.",
    "A cat is sitting on a bookshelf surrounded by books.",
    "A cat is grooming itself on a soft blanket.",
    "A cat is lying on a windowsill, eyes closed, in a quiet library.",
    "A dog sleeps peacefully on a cozy couch.",
    "A dog sniffs curiously at a flower garden.",
    "A dog pounces playfully at a butterfly in a meadow.",
    "A dog lounges lazily on a patio during a summer afternoon.",
    "A dog rests contentedly by a window with a view of the mountains.",
    "A tiger rests under the shade of a large tree in the jungle.",
    "A tiger slinks through the underbrush in the twilight.",
    "A tiger prowls the edge of a dense thicket, looking for prey.",
    "A tiger lounges in the shade, conserving energy for the hunt.",
    "A tiger rests under a waterfall, the cool water soothing its fur.",
    "A panda sleeps peacefully in a sunny spot in the forest.",
    "A panda munches on bamboo while sitting in a tree hollow.",
    "A panda walks along a path in a bamboo forest at dusk.",
    "A panda munches on bamboo while sitting on a log.",
    "A panda walks along a path in a bamboo forest at sunrise."
]

# Demo
image_path = "./Images/sharing.png"
text = "A cat is sleeping peacefully on a sunlit window sill."
clip_score = compute_image_text_alignment(image_path, text)
print(clip_score)

# Experiment 1
# excel_path = '/root/dataDisk/Data/Data1.xlsx'
# base_folder_path = '/root/dataDisk/Experiment1/seed2024/personal_prompt'
# folder_paths = [f"{base_folder_path}{str(i)}/" for i in range(20)]
# for index, folder_path in enumerate(folder_paths):
#     print(index)
#     process_images_and_update_excel(folder_path, excel_path, personal_prompts)

# Experiment 2
# excel_path = '/root/dataDisk/Data/Data2.xlsx'
# base_folder_path = '/root/dataDisk/Experiment2/'
# folder_paths = [f"{base_folder_path}seed{str(i)}/public_prompt0/" for i in range(1,6)]
# for index, folder_path in enumerate(folder_paths):
#     print(index)
#     process_images_and_update_excel(folder_path, excel_path, personal_prompts)

# Experiment 3
# excel_path = '/root/dataDisk/Data/Data3.xlsx'
# base_folder_path = '/root/dataDisk/Experiment3/'
# folder_paths = [f"{base_folder_path}seed{str(i)}/public_prompt1/" for i in range(1,6)]
# for index, folder_path in enumerate(folder_paths):
#     print(index)
#     process_images_and_update_excel(folder_path, excel_path, personal_prompts)

# Experiment 4
# excel_path = '/root/dataDisk/Data/Data4.xlsx'
# base_folder_path = '/root/dataDisk/Experiment4/'
# folder_paths = [f"{base_folder_path}seed{str(i)}/public_prompt2/" for i in range(1,6)]
# for index, folder_path in enumerate(folder_paths):
#     print(index)
#     process_images_and_update_excel(folder_path, excel_path, personal_prompts)

# Experiment 5
# excel_path = '/root/dataDisk/Data/Data5.xlsx'
# base_folder_path = '/root/dataDisk/Experiment5/'
# folder_paths = [f"{base_folder_path}seed{str(i)}/public_prompt3/" for i in range(1,6)]
# for index, folder_path in enumerate(folder_paths):
#     print(index)
#     process_images_and_update_excel(folder_path, excel_path, personal_prompts)
